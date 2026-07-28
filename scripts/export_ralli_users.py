"""Export the sanitized Ralli user directory for the static dashboard.

Only the operational directory sheet (Trang_tính3) is read. The hidden
credential sheet and its password column are intentionally never accessed.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "TLA Ralli.xlsx"
OUTPUT = ROOT / "ralli-users.js"
SOURCE_SHEET_INDEX = 1


def clean(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def account_type(login: str, full_name: str) -> str:
    normalized_login = login.casefold()
    normalized_name = full_name.casefold()
    if ".xemdon" in normalized_login or ".taodon" in normalized_login:
        return "service"
    if normalized_login.split(".")[-1].startswith("pgd") or normalized_name.startswith("pgd"):
        return "service"
    return "person"


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.worksheets[SOURCE_SHEET_INDEX]
    rows = sheet.iter_rows(min_row=2, values_only=True)
    users: dict[str, dict[str, str]] = {}

    for row in rows:
        department = clean(row[1])
        login = clean(row[2]).casefold()
        full_name = clean(row[3])
        email = clean(row[4]).casefold()
        role = clean(row[5])
        status = clean(row[6])
        if not login:
            continue

        record = {
            "login": login,
            "name": full_name,
            "email": email,
            "department": department,
            "role": role,
            "status": status,
            "accountType": account_type(login, full_name),
        }
        previous = users.get(login)
        if previous is not None and previous != record:
            raise ValueError(f"Conflicting duplicate login: {login}")
        users[login] = record

    result = sorted(
        users.values(),
        key=lambda user: (user["department"].casefold(), user["login"]),
    )
    if len(result) != 622:
        raise ValueError(f"Expected 622 unique Ralli users, got {len(result)}")
    if any(not user["name"] or not user["email"] or not user["department"] for user in result):
        raise ValueError("A Ralli user is missing name, email or department")

    payload = json.dumps(result, ensure_ascii=False, separators=(",", ":"))
    content = (
        "/* Generated from data/TLA Ralli.xlsx / Trang_tính3.\n"
        "   Sanitized identity and organization fields only. */\n"
        f"window.RALLI_USERS={payload};\n"
    )
    OUTPUT.write_text(content, encoding="utf-8", newline="\n")
    print(f"Exported {len(result)} Ralli users to {OUTPUT.name}")


if __name__ == "__main__":
    main()
